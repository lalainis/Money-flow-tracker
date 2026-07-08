import io
import uuid
from datetime import UTC, date, datetime, timedelta

from openpyxl import load_workbook

import settings


def auth_headers(token):
    return {"Authorization": f"Bearer {token}"}


def create_member(client, token, **overrides):
    payload = {
        "first_name": "Janis",
        "last_name": "Berzins",
        "phone": "20000001",
        "status": "Member",
        "membership_fee": 12,
        "joining_fee_paid": False,
        "role": "member",
    }
    payload.update(overrides)
    return client.post("/api/members", json=payload, headers=auth_headers(token))


def test_auth_init_rejects_invalid_phone(client):
    response = client.post("/api/auth/init", json={"phone": "abc"})

    assert response.status_code == 400
    assert "8 digits" in response.get_json()["error"]


def test_admin_login_returns_token_and_role(client):
    response = client.post("/api/auth/login", json={"phone": "29123456", "pin": "0308"})

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["token"]
    assert payload["user"]["role"] == "admin"


def test_expired_token_is_rejected(client, app_module):
    with app_module.app.app_context():
        admin = app_module.Member.query.filter_by(phone="29123456").first()
        assert admin is not None
        expired_token = str(uuid.uuid4())
        app_module.db.session.add(
            app_module.AuthToken(
                token=expired_token,
                member_id=admin.id,
                expires_at=datetime.now(UTC) - timedelta(minutes=1),
            )
        )
        app_module.db.session.commit()

    response = client.get("/api/dashboard", headers=auth_headers(expired_token))
    assert response.status_code == 401


def test_login_locks_after_repeated_failed_pin_attempts(client):
    for _ in range(4):
        wrong = client.post("/api/auth/login", json={"phone": "29123456", "pin": "9999"})
        assert wrong.status_code == 401

    fifth = client.post("/api/auth/login", json={"phone": "29123456", "pin": "9999"})
    assert fifth.status_code == 429

    blocked = client.post("/api/auth/login", json={"phone": "29123456", "pin": "0308"})
    assert blocked.status_code == 429


def test_non_admin_cannot_create_admin_account(client, admin_token):
    create_board = create_member(
        client,
        admin_token,
        phone="20000002",
        role="board",
        status="Biedrs",
    )
    assert create_board.status_code == 201

    setup_pin = client.post(
        "/api/auth/setup-pin",
        json={"phone": "20000002", "pin": "1234", "pin_confirm": "1234"},
    )
    assert setup_pin.status_code == 200

    board_login = client.post(
        "/api/auth/login",
        json={"phone": "20000002", "pin": "1234"},
    )
    assert board_login.status_code == 200
    board_token = board_login.get_json()["token"]

    create_admin = create_member(
        client,
        board_token,
        phone="20000003",
        role="admin",
        status="Biedrs",
    )

    assert create_admin.status_code == 403


def test_members_list_hides_admin_for_non_admin(client, admin_token):
    create_board = create_member(
        client,
        admin_token,
        phone="20000004",
        role="board",
        status="Biedrs",
    )
    assert create_board.status_code == 201

    setup_pin = client.post(
        "/api/auth/setup-pin",
        json={"phone": "20000004", "pin": "2222", "pin_confirm": "2222"},
    )
    assert setup_pin.status_code == 200

    board_login = client.post(
        "/api/auth/login",
        json={"phone": "20000004", "pin": "2222"},
    )
    assert board_login.status_code == 200
    board_token = board_login.get_json()["token"]

    response = client.get("/api/members", headers=auth_headers(board_token))

    assert response.status_code == 200
    phones = {row["phone"] for row in response.get_json()}
    assert "29123456" not in phones


def test_updating_period_resets_joining_fee_on_new_season(client, admin_token, app_module):
    created = create_member(
        client,
        admin_token,
        phone="20000005",
        status="Biedrs",
        joining_fee_paid=True,
        membership_fee=10,
    )
    assert created.status_code == 201
    member_id = created.get_json()["id"]

    first_update = client.post(
        "/api/period",
        json={
            "season_label": "2025/2026",
            "default_membership_fee": 30,
            "carry_over": 0,
        },
        headers=auth_headers(admin_token),
    )
    assert first_update.status_code == 200

    second_update = client.post(
        "/api/period",
        json={
            "season_label": "2026/2027",
            "default_membership_fee": 30,
            "carry_over": 0,
        },
        headers=auth_headers(admin_token),
    )
    assert second_update.status_code == 200

    with app_module.app.app_context():
        member = app_module.db.session.get(app_module.Member, member_id)
        assert member is not None
        assert member.joining_fee_paid is False


def test_add_other_income_rejects_non_positive_amount(client, admin_token):
    response = client.post(
        "/api/incomes",
        json={"amount": 0, "entry_date": date.today().isoformat(), "description": "x"},
        headers=auth_headers(admin_token),
    )

    assert response.status_code == 400
    assert "greater than 0" in response.get_json()["error"]


def test_expense_upload_and_attachment_download(client, admin_token, app_module):
    response = client.post(
        "/api/expenses",
        data={
            "category": "Citi",
            "amount": "15.50",
            "entry_date": date.today().isoformat(),
            "description": "Tests ar pielikumu",
            "attachment": (io.BytesIO(b"kvits"), "receipt.txt"),
        },
        content_type="multipart/form-data",
        headers=auth_headers(admin_token),
    )
    assert response.status_code == 201

    with app_module.app.app_context():
        expense = app_module.Expense.query.order_by(app_module.Expense.id.desc()).first()
        assert expense is not None
        assert expense.attachment
        saved_file = app_module.UPLOAD_DIR / expense.attachment
        assert saved_file.exists()
        attachment_name = expense.attachment

    download = client.get(f"/api/attachments/{attachment_name}", headers=auth_headers(admin_token))
    assert download.status_code == 200
    assert download.data == b"kvits"


def test_expense_upload_rejects_unsupported_attachment_type(client, admin_token):
    response = client.post(
        "/api/expenses",
        data={
            "category": "Citi",
            "amount": "4.50",
            "entry_date": date.today().isoformat(),
            "description": "Slikts pielikums",
            "attachment": (io.BytesIO(b"binary"), "malware.exe"),
        },
        content_type="multipart/form-data",
        headers=auth_headers(admin_token),
    )

    assert response.status_code == 400
    assert "Invalid file type" in response.get_json()["error"]


def test_expense_upload_rejects_malformed_multipart(client, admin_token):
    response = client.post(
        "/api/expenses",
        data=b"not-a-valid-multipart-body",
        content_type="multipart/form-data; boundary=broken",
        headers=auth_headers(admin_token),
    )

    assert response.status_code == 400
    assert "Invalid multipart request" in response.get_json()["error"]


def test_member_role_cannot_add_expense(client, admin_token, app_module):
    created = create_member(
        client,
        admin_token,
        phone="20000009",
        role="member",
        status="Biedrs",
    )
    assert created.status_code == 201
    member_id = created.get_json()["id"]

    with app_module.app.app_context():
        token = str(uuid.uuid4())
        app_module.db.session.add(
            app_module.AuthToken(
                token=token,
                member_id=member_id,
                expires_at=datetime.now(UTC) + timedelta(hours=1),
            )
        )
        app_module.db.session.commit()

    response = client.post(
        "/api/expenses",
        data={
            "category": "Citi",
            "amount": "10.00",
            "entry_date": date.today().isoformat(),
            "description": "Member mēģina pievienot izdevumu",
        },
        content_type="multipart/form-data",
        headers=auth_headers(token),
    )

    assert response.status_code == 403


def test_attachment_access_limited_to_owner_board_admin_and_auditor(client, admin_token, app_module):
    cashier_owner = create_member(
        client,
        admin_token,
        phone="20000010",
        role="cashier",
        status="Biedrs",
    )
    assert cashier_owner.status_code == 201

    cashier_other = create_member(
        client,
        admin_token,
        phone="20000011",
        role="cashier",
        status="Biedrs",
    )
    assert cashier_other.status_code == 201

    board_member = create_member(
        client,
        admin_token,
        phone="20000012",
        role="board",
        status="Biedrs",
    )
    assert board_member.status_code == 201

    auditor_member = create_member(
        client,
        admin_token,
        phone="20000013",
        role="auditor",
        status="Biedrs",
    )
    assert auditor_member.status_code == 201

    for phone, pin in [("20000010", "1010"), ("20000011", "1111"), ("20000012", "1212"), ("20000013", "1313")]:
        setup = client.post(
            "/api/auth/setup-pin",
            json={"phone": phone, "pin": pin, "pin_confirm": pin},
        )
        assert setup.status_code == 200

    owner_login = client.post("/api/auth/login", json={"phone": "20000010", "pin": "1010"})
    other_login = client.post("/api/auth/login", json={"phone": "20000011", "pin": "1111"})
    board_login = client.post("/api/auth/login", json={"phone": "20000012", "pin": "1212"})
    auditor_login = client.post("/api/auth/login", json={"phone": "20000013", "pin": "1313"})
    assert owner_login.status_code == 200
    assert other_login.status_code == 200
    assert board_login.status_code == 200
    assert auditor_login.status_code == 200

    owner_token = owner_login.get_json()["token"]
    other_token = other_login.get_json()["token"]
    board_token = board_login.get_json()["token"]
    auditor_token = auditor_login.get_json()["token"]

    create_expense = client.post(
        "/api/expenses",
        data={
            "category": "Citi",
            "amount": "18.00",
            "entry_date": date.today().isoformat(),
            "description": "Īpašnieka pielikums",
            "attachment": (io.BytesIO(b"only-owner-board-admin"), "owner.txt"),
        },
        content_type="multipart/form-data",
        headers=auth_headers(owner_token),
    )
    assert create_expense.status_code == 201

    with app_module.app.app_context():
        expense = app_module.Expense.query.order_by(app_module.Expense.id.desc()).first()
        assert expense is not None
        attachment_name = expense.attachment

    owner_download = client.get(f"/api/attachments/{attachment_name}", headers=auth_headers(owner_token))
    assert owner_download.status_code == 200

    other_download = client.get(f"/api/attachments/{attachment_name}", headers=auth_headers(other_token))
    assert other_download.status_code == 403

    board_download = client.get(f"/api/attachments/{attachment_name}", headers=auth_headers(board_token))
    assert board_download.status_code == 200

    auditor_download = client.get(f"/api/attachments/{attachment_name}", headers=auth_headers(auditor_token))
    assert auditor_download.status_code == 200

    admin_download = client.get(f"/api/attachments/{attachment_name}", headers=auth_headers(admin_token))
    assert admin_download.status_code == 200


def test_audit_logs_forbidden_for_board_role(client, admin_token):
    create_board = create_member(
        client,
        admin_token,
        phone="20000006",
        role="board",
        status="Biedrs",
    )
    assert create_board.status_code == 201

    setup_pin = client.post(
        "/api/auth/setup-pin",
        json={"phone": "20000006", "pin": "3333", "pin_confirm": "3333"},
    )
    assert setup_pin.status_code == 200

    board_login = client.post(
        "/api/auth/login",
        json={"phone": "20000006", "pin": "3333"},
    )
    assert board_login.status_code == 200
    board_token = board_login.get_json()["token"]

    response = client.get("/api/audit-logs", headers=auth_headers(board_token))
    assert response.status_code == 403


def test_audit_logs_require_auth(client):
    response = client.get("/api/audit-logs")
    assert response.status_code == 401


def test_admin_sees_created_audit_log_entries(client, admin_token, app_module):
    income_create = client.post(
        "/api/incomes",
        json={"amount": 10, "entry_date": date.today().isoformat(), "description": "pirms"},
        headers=auth_headers(admin_token),
    )
    assert income_create.status_code == 201

    with app_module.app.app_context():
        income = app_module.Income.query.order_by(app_module.Income.id.desc()).first()
        assert income is not None
        income_id = income.id

    income_update = client.put(
        f"/api/incomes/{income_id}",
        json={"amount": 11, "entry_date": date.today().isoformat(), "description": "pec"},
        headers=auth_headers(admin_token),
    )
    assert income_update.status_code == 200

    logs_response = client.get("/api/audit-logs", headers=auth_headers(admin_token))
    assert logs_response.status_code == 200
    logs = logs_response.get_json()["audit_logs"]
    assert any(log["entity_type"] == "income" and log["action"] == "update" for log in logs)


def test_board_cannot_update_period_after_first_save(client, admin_token):
    create_board = create_member(
        client,
        admin_token,
        phone="20000007",
        role="board",
        status="Biedrs",
    )
    assert create_board.status_code == 201

    setup_pin = client.post(
        "/api/auth/setup-pin",
        json={"phone": "20000007", "pin": "4444", "pin_confirm": "4444"},
    )
    assert setup_pin.status_code == 200

    board_login = client.post(
        "/api/auth/login",
        json={"phone": "20000007", "pin": "4444"},
    )
    assert board_login.status_code == 200
    board_token = board_login.get_json()["token"]

    first_update = client.post(
        "/api/period",
        json={
            "season_label": "2027/2028",
            "default_membership_fee": 25,
            "carry_over": 100,
        },
        headers=auth_headers(board_token),
    )
    assert first_update.status_code == 200

    second_update = client.post(
        "/api/period",
        json={
            "season_label": "2027/2028",
            "default_membership_fee": 30,
            "carry_over": 100,
        },
        headers=auth_headers(board_token),
    )

    assert second_update.status_code == 403
    assert "only by admin" in second_update.get_json()["error"]


def test_board_cannot_change_carry_over_after_first_save(client, admin_token):
    create_board = create_member(
        client,
        admin_token,
        phone="20000008",
        role="board",
        status="Biedrs",
    )
    assert create_board.status_code == 201

    setup_pin = client.post(
        "/api/auth/setup-pin",
        json={"phone": "20000008", "pin": "5555", "pin_confirm": "5555"},
    )
    assert setup_pin.status_code == 200

    board_login = client.post(
        "/api/auth/login",
        json={"phone": "20000008", "pin": "5555"},
    )
    assert board_login.status_code == 200
    board_token = board_login.get_json()["token"]

    first_update = client.post(
        "/api/period",
        json={
            "season_label": "2028/2029",
            "default_membership_fee": 0,
            "carry_over": 10,
        },
        headers=auth_headers(board_token),
    )
    assert first_update.status_code == 200

    second_update = client.post(
        "/api/period",
        json={
            "season_label": "2028/2029",
            "default_membership_fee": 0,
            "carry_over": 20,
        },
        headers=auth_headers(board_token),
    )

    assert second_update.status_code == 403
    assert "carry over" in second_update.get_json()["error"]


def test_export_contains_expected_sheets_and_summary(client, admin_token):
    income_create = client.post(
        "/api/incomes",
        json={"amount": 50, "entry_date": date.today().isoformat(), "description": "Eksporta tests"},
        headers=auth_headers(admin_token),
    )
    assert income_create.status_code == 201

    expense_create = client.post(
        "/api/expenses",
        data={
            "category": "Citi",
            "amount": "20.00",
            "entry_date": date.today().isoformat(),
            "description": "Eksporta izdevumi",
        },
        content_type="multipart/form-data",
        headers=auth_headers(admin_token),
    )
    assert expense_create.status_code == 201

    export_response = client.get("/api/export", headers=auth_headers(admin_token))
    assert export_response.status_code == 200

    workbook = load_workbook(io.BytesIO(export_response.data))
    assert set(["Incomes", "Expenses", "Summary", "Members", "Audit log"]).issubset(workbook.sheetnames)

    summary_sheet = workbook["Summary"]
    assert summary_sheet["A1"].value == "Reporting period"
    assert summary_sheet["A2"].value == "Incomes EUR"
    assert summary_sheet["A3"].value == "Expenses EUR"


def test_login_sets_cookie_and_cookie_auth_works(client):
    login = client.post("/api/auth/login", json={"phone": "29123456", "pin": "0308"})
    assert login.status_code == 200
    set_cookie_header = login.headers.get("Set-Cookie", "")
    assert "auth_token=" in set_cookie_header
    assert "HttpOnly" in set_cookie_header

    dashboard = client.get("/api/dashboard")
    assert dashboard.status_code == 200


def test_auth_init_rate_limit_per_phone(client):
    # AUTH_INIT_RATE_PHONE is a Flask-Limiter rate string (for example, "10 per minute").
    limit_count = int(str(settings.AUTH_INIT_RATE_PHONE).split(" ", 1)[0])

    for _ in range(limit_count):
        response = client.post("/api/auth/init", json={"phone": "29123456"})
        assert response.status_code == 200

    limited = client.post("/api/auth/init", json={"phone": "29123456"})
    assert limited.status_code == 429
