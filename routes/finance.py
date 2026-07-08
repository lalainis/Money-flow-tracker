import json
import uuid
from io import BytesIO
from datetime import date

from flask import Blueprint, jsonify, request, send_file, send_from_directory
from openpyxl import Workbook
from werkzeug.utils import secure_filename

from app_core import db
from models import AuditLog, Expense, Income, Member, MemberSeasonFee
from services import (
    create_audit_log,
    current_period,
    expense_to_payload,
    get_period_for_request,
    income_to_payload,
    is_allowed_attachment,
    localize_income_type,
    normalize_expense_category,
    normalize_role,
    parse_entry_date,
    period_totals,
    recalculate_member_paid_this_period,
    season_label_for_date,
    to_decimal,
    token_required,
)
from settings import EXPENSE_CATEGORIES, UPLOAD_DIR

finance_bp = Blueprint("finance", __name__)


@finance_bp.route("/api/incomes", methods=["POST"])
@token_required({"cashier", "admin"})
def add_other_income():
    data = request.get_json() or {}
    try:
        amount = to_decimal(data.get("amount", 0) or 0)
        entry_date = parse_entry_date(data.get("entry_date", date.today().isoformat()))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    description = str(data.get("description", "")).strip()

    if amount <= 0:
        return jsonify({"error": "Amount must be greater than 0"}), 400

    income = Income(
        income_type="other_income",
        amount=amount,
        entry_date=entry_date,
        description=description,
    )
    db.session.add(income)
    db.session.commit()
    return jsonify({"message": "Income added"}), 201


@finance_bp.route("/api/expenses", methods=["POST"])
@token_required({"cashier", "admin"})
def add_expense():
    if request.mimetype == "multipart/form-data" and not request.form and not request.files:
        return jsonify({"error": "Invalid multipart request"}), 400

    category = normalize_expense_category(request.form.get("category"))
    amount_raw = request.form.get("amount", "0")
    entry_date_raw = request.form.get("entry_date", date.today().isoformat())
    description = (request.form.get("description") or "").strip()

    if category not in EXPENSE_CATEGORIES:
        return jsonify({"error": "Invalid expense category"}), 400

    try:
        amount = to_decimal(amount_raw)
        entry_date = parse_entry_date(entry_date_raw)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    if amount <= 0:
        return jsonify({"error": "Amount must be greater than 0"}), 400

    attachment_name = None
    file = request.files.get("attachment")
    if file and file.filename:
        safe_name = secure_filename(file.filename)
        if not safe_name or not is_allowed_attachment(safe_name):
            return jsonify({"error": "Invalid file type"}), 400
        unique_name = f"{uuid.uuid4()}_{safe_name}"
        file.save(UPLOAD_DIR / unique_name)
        attachment_name = unique_name

    expense = Expense(
        category=category,
        amount=amount,
        entry_date=entry_date,
        description=description,
        attachment=attachment_name,
        created_by_member_id=request.current_user.id,
    )

    db.session.add(expense)
    db.session.commit()

    return jsonify({"message": "Expense added"}), 201


@finance_bp.route("/api/history")
@token_required({"cashier", "board", "admin", "auditor"})
def history():
    period = get_period_for_request()
    requester_role = normalize_role(request.current_user.role)

    incomes = (
        Income.query.filter(
            Income.entry_date >= period.start_date,
            Income.entry_date <= period.end_date,
        )
        .order_by(Income.entry_date.desc())
        .all()
    )

    expenses = (
        Expense.query.filter(
            Expense.entry_date >= period.start_date,
            Expense.entry_date <= period.end_date,
        )
        .order_by(Expense.entry_date.desc())
        .all()
    )

    income_rows = []
    for row in incomes:
        member = db.session.get(Member, row.member_id) if row.member_id else None
        member_name = ""
        if member and (requester_role == "admin" or normalize_role(member.role) != "admin"):
            member_name = f"{member.first_name} {member.last_name}"
        income_rows.append(
            {
                "id": row.id,
                "type": localize_income_type(row.income_type),
                "member_name": member_name,
                "amount": float(row.amount),
                "entry_date": row.entry_date.isoformat(),
                "description": row.description or "",
            }
        )

    expense_rows = []
    for row in expenses:
        expense_rows.append(
            {
                "id": row.id,
                "category": row.category,
                "amount": float(row.amount),
                "entry_date": row.entry_date.isoformat(),
                "description": row.description or "",
                "attachment": row.attachment,
            }
        )

    return jsonify({"incomes": income_rows, "expenses": expense_rows})


@finance_bp.route("/api/incomes/<int:income_id>", methods=["PUT"])
@token_required({"admin"})
def update_income(income_id):
    income = db.session.get(Income, income_id)
    if not income:
        return jsonify({"error": "Income record not found"}), 404

    data = request.get_json() or {}
    try:
        amount = to_decimal(data.get("amount", income.amount) or 0)
        entry_date = parse_entry_date(data.get("entry_date", income.entry_date.isoformat()))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    if amount <= 0:
        return jsonify({"error": "Amount must be greater than 0"}), 400

    previous_member_fee = income.income_type in {"member_fee", "biedra nauda"}
    previous_member_id = income.member_id
    old_payload = income_to_payload(income)

    income.amount = amount
    income.entry_date = entry_date
    income.description = str(data.get("description", income.description or "")).strip()

    if previous_member_fee:
        recalculate_member_paid_this_period(previous_member_id)

    create_audit_log(
        action="update",
        entity_type="income",
        entity_id=income.id,
        season_label=season_label_for_date(income.entry_date),
        old_data=old_payload,
        new_data=income_to_payload(income),
    )

    db.session.commit()
    return jsonify({"message": "Income record updated"})


@finance_bp.route("/api/expenses/<int:expense_id>", methods=["PUT"])
@token_required({"admin"})
def update_expense(expense_id):
    expense = db.session.get(Expense, expense_id)
    if not expense:
        return jsonify({"error": "Expense record not found"}), 404

    data = request.get_json() or {}
    old_payload = expense_to_payload(expense)
    category = normalize_expense_category(data.get("category", expense.category))
    if category not in EXPENSE_CATEGORIES:
        return jsonify({"error": "Invalid expense category"}), 400

    try:
        amount = to_decimal(data.get("amount", expense.amount) or 0)
        entry_date = parse_entry_date(data.get("entry_date", expense.entry_date.isoformat()))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    if amount <= 0:
        return jsonify({"error": "Amount must be greater than 0"}), 400

    expense.category = category
    expense.amount = amount
    expense.entry_date = entry_date
    expense.description = str(data.get("description", expense.description or "")).strip()

    create_audit_log(
        action="update",
        entity_type="expense",
        entity_id=expense.id,
        season_label=season_label_for_date(expense.entry_date),
        old_data=old_payload,
        new_data=expense_to_payload(expense),
    )

    db.session.commit()
    return jsonify({"message": "Expense record updated"})


@finance_bp.route("/api/incomes/<int:income_id>", methods=["DELETE"])
@token_required({"admin"})
def delete_income(income_id):
    income = db.session.get(Income, income_id)
    if not income:
        return jsonify({"error": "Income record not found"}), 404

    is_member_fee = income.income_type in {"member_fee", "biedra nauda"}
    member_id = income.member_id
    season_label = season_label_for_date(income.entry_date)
    old_payload = income_to_payload(income)

    db.session.delete(income)
    if is_member_fee:
        recalculate_member_paid_this_period(member_id)

    create_audit_log(
        action="delete",
        entity_type="income",
        entity_id=income_id,
        season_label=season_label,
        old_data=old_payload,
        new_data=None,
    )

    db.session.commit()
    return jsonify({"message": "Income record deleted"})


@finance_bp.route("/api/expenses/<int:expense_id>", methods=["DELETE"])
@token_required({"admin"})
def delete_expense(expense_id):
    expense = db.session.get(Expense, expense_id)
    if not expense:
        return jsonify({"error": "Expense record not found"}), 404

    season_label = season_label_for_date(expense.entry_date)
    old_payload = expense_to_payload(expense)
    db.session.delete(expense)

    create_audit_log(
        action="delete",
        entity_type="expense",
        entity_id=expense_id,
        season_label=season_label,
        old_data=old_payload,
        new_data=None,
    )

    db.session.commit()
    return jsonify({"message": "Expense record deleted"})


@finance_bp.route("/api/audit-logs")
@token_required({"admin"})
def list_audit_logs():
    season_label = (request.args.get("season_label") or "").strip()
    query = AuditLog.query
    if season_label:
        query = query.filter(AuditLog.season_label == season_label)

    rows = query.order_by(AuditLog.changed_at.desc()).limit(300).all()
    result = []
    for row in rows:
        actor = db.session.get(Member, row.changed_by_member_id)
        actor_name = ""
        if actor:
            actor_name = f"{actor.first_name} {actor.last_name}"
        result.append(
            {
                "id": row.id,
                "action": row.action,
                "entity_type": row.entity_type,
                "entity_id": row.entity_id,
                "season_label": row.season_label,
                "changed_by": actor_name,
                "changed_at": row.changed_at.isoformat(),
                "old_data": json.loads(row.old_data) if row.old_data else None,
                "new_data": json.loads(row.new_data) if row.new_data else None,
            }
        )

    return jsonify({"audit_logs": result})


@finance_bp.route("/api/attachments/<path:filename>")
@token_required()
def get_attachment(filename):
    expense = Expense.query.filter_by(attachment=filename).first()
    if not expense:
        return jsonify({"error": "Attachment not found"}), 404

    requester_role = normalize_role(request.current_user.role)
    is_owner = expense.created_by_member_id == request.current_user.id
    if requester_role not in {"board", "admin", "auditor"} and not is_owner:
        return jsonify({"error": "Insufficient permissions"}), 403

    return send_from_directory(UPLOAD_DIR, filename, as_attachment=False)


@finance_bp.route("/api/export")
@token_required({"cashier", "board", "admin", "auditor"})
def export_balance():
    period = get_period_for_request()
    requester_role = normalize_role(request.current_user.role)
    totals = period_totals(period)

    incomes = (
        Income.query.filter(
            Income.entry_date >= period.start_date,
            Income.entry_date <= period.end_date,
        )
        .order_by(Income.entry_date.asc())
        .all()
    )

    expenses = (
        Expense.query.filter(
            Expense.entry_date >= period.start_date,
            Expense.entry_date <= period.end_date,
        )
        .order_by(Expense.entry_date.asc())
        .all()
    )

    wb = Workbook()
    ws_income = wb.active
    ws_income.title = "Incomes"

    ws_income.append(["Type", "Member", "Amount EUR", "Date", "Description"])
    ws_income.append(["Carry over from previous period", "", float(period.carry_over), "", ""])

    for row in incomes:
        member = db.session.get(Member, row.member_id) if row.member_id else None
        member_name = ""
        if member and (requester_role == "admin" or normalize_role(member.role) != "admin"):
            member_name = f"{member.first_name} {member.last_name}"
        ws_income.append(
            [
                localize_income_type(row.income_type),
                member_name,
                float(row.amount),
                row.entry_date.isoformat(),
                row.description or "",
            ]
        )

    ws_income.append(["Total incomes EUR", "", totals["income_total"], "", ""])

    ws_expense = wb.create_sheet("Expenses")
    ws_expense.append(["Category", "Amount EUR", "Date", "Description", "Attachment"])

    for row in expenses:
        ws_expense.append(
            [
                row.category,
                float(row.amount),
                row.entry_date.isoformat(),
                row.description or "",
                row.attachment or "",
            ]
        )

    ws_expense.append(["Total expenses EUR", totals["expense_total"], "", "", ""])

    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Reporting period", period.season_label])
    ws_summary.append(["Incomes EUR", totals["income_total"]])
    ws_summary.append(["Expenses EUR", totals["expense_total"]])
    if totals["difference"] >= 0:
        ws_summary.append(["Balance EUR", totals["balance"]])
    else:
        ws_summary.append(["Deficit EUR", totals["deficit"]])

    ws_members = wb.create_sheet("Members")
    ws_members.append(["No.", "First name", "Last name", "Status", "Due EUR", "Paid EUR", "Joining fee"])
    all_members = Member.query.order_by(Member.list_no).all()
    if requester_role != "admin":
        all_members = [m for m in all_members if normalize_role(m.role) != "admin"]
    paid_rows_map = {
        mid: tot
        for mid, tot in (
            db.session.query(Income.member_id, db.func.coalesce(db.func.sum(Income.amount), 0))
            .filter(
                Income.income_type.in_(["member_fee", "biedra nauda"]),
                Income.member_id.isnot(None),
                Income.entry_date >= period.start_date,
                Income.entry_date <= period.end_date,
            )
            .group_by(Income.member_id)
            .all()
        )
    }
    fee_rows_map = {
        row.member_id: row.membership_fee
        for row in MemberSeasonFee.query.filter_by(season_label=period.season_label).all()
    }
    for m in all_members:
        fee = float(fee_rows_map.get(m.id, m.membership_fee))
        paid = float(paid_rows_map.get(m.id, 0))
        ws_members.append([m.list_no, m.first_name, m.last_name, m.status, fee, paid, "Yes" if m.joining_fee_paid else "No"])

    ws_audit = wb.create_sheet("Audit log")
    ws_audit.append(["Time", "Action", "Record type", "Record ID", "User", "Old value", "New value"])
    audit_rows = AuditLog.query.filter_by(season_label=period.season_label).order_by(AuditLog.changed_at.asc()).all()
    for row in audit_rows:
        actor = db.session.get(Member, row.changed_by_member_id)
        actor_name = ""
        if actor:
            actor_name = f"{actor.first_name} {actor.last_name}"
        ws_audit.append(
            [
                row.changed_at.isoformat(sep=" ", timespec="seconds"),
                "Updated" if row.action == "update" else "Deleted",
                "Income" if row.entity_type == "income" else "Expense",
                row.entity_id,
                actor_name,
                row.old_data or "",
                row.new_data or "",
            ]
        )

    filename = f"bilance_{period.season_label.replace('/', '-')}.xlsx"
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
